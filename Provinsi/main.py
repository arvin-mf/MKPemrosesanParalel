import time
import requests
from  SubdistrictReq.subdistrict_req import SubdistrictRequests
# from  SubdistrictReq.subdistrict_req import subdistrictsCount
from  WeatherReq.weather_req import WeatherRequests
from openpyxl import load_workbook

workbook = load_workbook('C:/Users/Public/Avrin/Provinsi/Gorontalo.xlsx')
sheet = workbook.active

codes = []
names = []

def getDistricts(id):
    kota = id
    key = '[api_key]'
    api = 'https://api.goapi.io/regional/kota'
    payload = {
        "provinsi_id": kota,
        "api_key": key
    }
    response = requests.get(api, params=payload, timeout=1000)

    global i
    i = 0
    for r in response.json()['data']:
        codes.append(r['id'])

# def getCoordinate(name, row):
#     api = 'https://montanaflynn-geocoder.p.rapidapi.com/address'
#     key = '[api_key]'
#     city = name
#     payload = {
#         "address": city,
#         "X-RapidAPI-Key": key,
# 	      "X-RapidAPI-Host": "montanaflynn-geocoder.p.rapidapi.com"
#     }
#     response = requests.get(api, params=payload, timeout=1000)
#     sheet.cell(column=4, row=row, value=f"{response.json()['latitude']},{response.json()['longitude']}")
#     workbook.save('C:/Users/Public/Avrin/Provinsi/Gorontalo.xlsx')


def main():
    getDistricts('75')      # 75: id prov.Gorontalo

    start_time = time.time()

    current_s_requests = []
    for c in codes:
        subdistrict_req = SubdistrictRequests(code=c, wb=workbook, ws=sheet, list=names)
        current_s_requests.append(subdistrict_req)

    for r in current_s_requests:
        r.join() 

    for n in range(77):
        sheet.cell(column=3, row=4+n, value=names[n])

    print("Getting subdistricts code took: {}".format(time.time() - start_time))

    start_time = time.time()

    count = 0
    current_w_requests = []
    for row in sheet["D4:D80"]:
        for cell in row:
            weather_req = WeatherRequests(city=cell.value, wb=workbook, ws=sheet, i=count)
            count += 1
            current_w_requests.append(weather_req)

    for r in current_w_requests:
        r.join()

    print("Getting weathers information took: {}".format(time.time() - start_time))


if __name__ == "__main__":
    main()
